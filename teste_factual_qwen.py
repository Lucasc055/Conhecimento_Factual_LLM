import re
import time
import unicodedata
from difflib import SequenceMatcher

import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# ── Configuração ──────────────────────────────────────────────────────────────

FILE_IN = "perguntas.xlsx"

MODEL_NAME = "qwen/qwen2.5-vl-7b"

SYSTEM_PROMPT = (
    "Respond using only established medical and scientific consensus. "
    "Be strictly factual. "
    "Return only one short affirmative answer in English. "
    "Do not explain. Do not use negation. "
    "Do not return sentences or lists. "
    "Return only the exact medical term, substance, condition, or technical term."
)

client = OpenAI(
    base_url="http://127.0.0.1:1234/v1",
    api_key="lm-studio"
)

# ── Funções auxiliares ────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    text = str(text).strip().lower()
    text = re.sub(r'^\d+[\.\)]\s*', '', text)

    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")

    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def contains_negation(text: str) -> bool:
    t = normalize(text)
    negation_patterns = [
        r'\bno\b', r'\bnot\b', r'\bnone\b', r'\bneither\b', r'\bnever\b',
        r'\bwithout\b', r'\bdoes not\b', r'\bdo not\b', r'\bdid not\b',
        r'\bis not\b', r'\bare not\b', r'\bwas not\b', r'\bwere not\b',
        r'\bcannot\b', r"\bcan't\b", r"\bdon't\b", r"\bdoesn't\b",
        r"\bisn't\b", r"\baren't\b", r"\bwasn't\b", r"\bweren't\b",
    ]
    return any(re.search(pattern, t) for pattern in negation_patterns)


def is_exact_match(target_raw: str, answer_raw: str) -> bool:
    return normalize(target_raw) == normalize(answer_raw)


def word_similarity(w1: str, w2: str) -> float:
    return SequenceMatcher(None, w1, w2).ratio()


def similarity_flag(target_raw: str, answer_raw: str) -> str:
    """
    Retorna:
    - EXATO
    - SEMELHANTE
    - DIFERENTE
    - VAZIO
    """
    t = normalize(target_raw)
    a = normalize(answer_raw)

    if not a:
        return "VAZIO"

    if t == a:
        return "EXATO"

    stopwords = {
        "the", "a", "an", "of", "and", "or", "is", "in", "to", "with",
        "for", "that", "this", "as", "are", "it", "its", "by", "on",
        "at", "from", "be", "than", "into", "such"
    }

    t_words = set(t.split()) - stopwords
    a_words = set(a.split()) - stopwords

    if not t_words or not a_words:
        return "DIFERENTE"

    if t_words & a_words:
        return "SEMELHANTE"

    for tw in t_words:
        for aw in a_words:
            if tw in aw or aw in tw:
                return "SEMELHANTE"

    for tw in t_words:
        for aw in a_words:
            if len(tw) >= 4 and len(aw) >= 4:
                if word_similarity(tw, aw) >= 0.82:
                    return "SEMELHANTE"

    if SequenceMatcher(None, t, a).ratio() >= 0.80:
        return "SEMELHANTE"

    return "DIFERENTE"


def clean_llm_output(raw: str) -> str:
    if raw is None:
        return ""

    text = str(raw).strip()
    text = re.sub(r'^\d+[\.\)]\s*', '', text).strip()
    text = text.strip(' "\'')

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if lines:
        text = lines[0]

    text = re.sub(r'[.;:,\s]+$', '', text).strip()
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def query_llm(question: str, retries: int = 2) -> str:
    for attempt in range(retries + 1):
        try:
            resp = client.chat.completions.create(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": question},
                ],
                temperature=0.3,
                max_tokens=60,
            )

            raw = resp.choices[0].message.content
            return clean_llm_output(raw)

        except Exception as e:
            if attempt < retries:
                time.sleep(2)
            else:
                print(f"   ⚠ Erro API (tentativa {attempt + 1}): {e}")
                return "ERRO_API"


def classify_answer(target: str, answer: str) -> tuple[str, str]:
    """
    Regras finais:
    - CORRETO: match exato
    - ERRADO: claramente diferente
    - REVISAO_MANUAL: apenas quando é semelhante
    - FALHA: erro técnico ou vazio
    """
    if answer == "ERRO_API" or not answer.strip():
        return "FALHA", "Erro API ou resposta vazia"

    if is_exact_match(target, answer):
        return "CORRETO", "Match exato"

    if contains_negation(answer):
        return "ERRADO", "Contém negação"

    flag = similarity_flag(target, answer)

    if flag == "SEMELHANTE":
        return "REVISAO_MANUAL", "SEMELHANTE"

    if flag == "DIFERENTE":
        return "ERRADO", "DIFERENTE"

    return "ERRADO", flag


def apply_excel_formatting(filepath: str):
    wb = load_workbook(filepath)
    ws = wb.active

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    header_fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    body_font = Font(name="Arial", size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value)

        if header == "Pergunta":
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 60
        elif header == "Resposta_Correta":
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 28
        elif header.startswith("LLM_R"):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 28
        elif header.startswith("Status_Automatico_R"):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22
        elif header.startswith("Observacao_R"):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22
        elif header.startswith("Validacao_Final_R"):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        elif header.startswith("Resposta_Final_R"):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 30
        else:
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    status_col = None
    final_col = None
    answer_final_col = None

    for name, idx in headers.items():
        if isinstance(name, str) and name.startswith("Status_Automatico_"):
            status_col = idx
        if isinstance(name, str) and name.startswith("Validacao_Final_"):
            final_col = idx
        if isinstance(name, str) and name.startswith("Resposta_Final_"):
            answer_final_col = idx

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if status_col:
            status_cell = ws.cell(row=row, column=status_col)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            if status_cell.value == "CORRETO":
                status_cell.fill = green
            elif status_cell.value == "ERRADO":
                status_cell.fill = red
            elif status_cell.value == "FALHA":
                status_cell.fill = orange
            elif status_cell.value == "REVISAO_MANUAL":
                status_cell.fill = blue

        if final_col:
            final_cell = ws.cell(row=row, column=final_col)
            final_cell.alignment = Alignment(horizontal="center", vertical="center")

            if final_cell.value == "CORRETO":
                final_cell.fill = green
            elif final_cell.value == "ERRADO":
                final_cell.fill = red

        if answer_final_col and final_col:
            answer_cell = ws.cell(row=row, column=answer_final_col)
            final_cell = ws.cell(row=row, column=final_col)

            if final_cell.value == "CORRETO":
                answer_cell.fill = green
            elif final_cell.value == "ERRADO":
                answer_cell.fill = red

    if final_col:
        col_letter = ws.cell(row=1, column=final_col).column_letter

        dv = DataValidation(
            type="list",
            formula1='"CORRETO,ERRADO"',
            allow_blank=True
        )
        dv.prompt = "Escolha CORRETO ou ERRADO"
        dv.error = "Valor inválido. Use apenas CORRETO ou ERRADO."
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    wb.save(filepath)


# ── Pipeline principal ────────────────────────────────────────────────────────

def run_audit():
    round_num = input("Número da execução (1, 2 ou 3): ").strip()
    if round_num not in {"1", "2", "3"}:
        print("❌ Introduza 1, 2 ou 3.")
        return

    model_slug = MODEL_NAME.replace("/", "_").replace(" ", "_")
    file_out = f"Relatorio_Factual_{model_slug}_R{round_num}.xlsx"

    try:
        df = pd.read_excel(FILE_IN)
    except FileNotFoundError:
        print(f"❌ Ficheiro '{FILE_IN}' não encontrado.")
        return

    required_cols = {"Pergunta", "Resposta_Correta"}
    if not required_cols.issubset(df.columns):
        print(f"❌ Colunas em falta: {required_cols - set(df.columns)}")
        return

    total = len(df)

    answers = []
    statuses = []
    notes = []
    final_validation = []
    final_answer_only = []

    correct_count = 0
    wrong_count = 0
    review_count = 0
    fail_count = 0

    print(f"\n🚀 Auditoria Factual — Modelo: {MODEL_NAME} | Rodada {round_num}")
    print(f"   {total} questões a processar...\n" + "─" * 90)

    for i, row in df.iterrows():
        question = str(row["Pergunta"]).strip()
        target = str(row["Resposta_Correta"]).strip()

        answer = query_llm(question)
        status, note = classify_answer(target, answer)

        if status == "CORRETO":
            correct_count += 1
            icon = "✅"
            final_value = "CORRETO"
            final_answer = answer
        elif status == "ERRADO":
            wrong_count += 1
            icon = "❌"
            final_value = "ERRADO"
            final_answer = answer
        elif status == "REVISAO_MANUAL":
            review_count += 1
            icon = "📝"
            final_value = ""
            final_answer = answer
        else:
            fail_count += 1
            icon = "⚠"
            final_value = ""
            final_answer = ""

        answers.append(answer)
        statuses.append(status)
        notes.append(note)
        final_validation.append(final_value)
        final_answer_only.append(final_answer)

        print(
            f"[{i + 1:02d}/{total}] {icon} {status:<16} | "
            f"Esperado: {target:<25} | LLM: {answer}"
        )

    df_out = df.copy()
    df_out[f"LLM_R{round_num}"] = answers
    df_out[f"Status_Automatico_R{round_num}"] = statuses
    df_out[f"Observacao_R{round_num}"] = notes
    df_out[f"Validacao_Final_R{round_num}"] = final_validation
    df_out[f"Resposta_Final_R{round_num}"] = final_answer_only

    df_out.to_excel(file_out, index=False)
    apply_excel_formatting(file_out)

    accuracy_auto = (correct_count / total) * 100 if total else 0

    print("─" * 90)
    print("✅ Concluído!")

if __name__ == "__main__":
    run_audit()